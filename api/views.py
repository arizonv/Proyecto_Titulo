from django.http import HttpResponse, JsonResponse
from django.core import serializers
from django.views import View
from django.contrib.auth import authenticate, login, logout
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from rest_framework.parsers import JSONParser
from rest_framework.exceptions import ParseError
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status, generics
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.authtoken.models import Token
from rest_framework.generics import ListAPIView


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook

from .serializers import (
    UserSerializer,
    LoginSerializer,
    listSerializer,
    createUserSerializer,
    ReservaSerializer,
)

from accounts.models import User, ComplejoDeportivo
from cliente.models import Cliente, Boleta, Reserva, Ticket

# Decorador
from accounts.decorators import has_permission



#API DE LOGIN PARA IBICIAR SESION 
class LoginAPIView(APIView):
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = authenticate(request, username=serializer.validated_data['username'], password=serializer.validated_data['password'])
        if user is not None:
            login(request, user)
            token, created = Token.objects.get_or_create(user=user)
            return Response({'id': user.id,
                             'usuario': user.username,
                             'admin': user.is_staff,
                             'token': token.key,
                             }, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

# API PARA CERRAR SESION
class UserLogout(APIView):
    authentication_classes = [TokenAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        logout(request)
        return Response('Logout successfully')

#API PARA LISTAR TODOS LOS USUARIOS
class UserList(generics.ListAPIView):
    # authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = User.objects.all()
    serializer_class = listSerializer
    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)

#API DE REGISTRO DE USUARIO ( POR DEFECTO COMO EXPLICA EN EL MODELO SE CREARA POR DEFECTO QUE EL USUARIO SEA CLIENTE YA QUE ES EL FORM DE REGISTRO DE LA PAGINA)
from rest_framework import generics
from rest_framework.response import Response
from .serializers import createUserSerializer, listSerializer

class Register(generics.GenericAPIView):
    serializer_class = createUserSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()

        roles_name = serializer.get_roles(user)
        user_data = listSerializer(user).data
        user_data['roles'] = roles_name

        return Response({
            "user": user_data,
            "message": "User Created Successfully.",
        })



#API PARA GENERAR REPORTES DE TODOS LOS MODELOS IMPORTANTES
# @method_decorator(has_permission(['']), name='dispatch')
class ExcelReportView(APIView):
    def generate_excel(self, data, headers, sheet_title, filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_title
        
        # Adding headers
        sheet.append(headers)
        
        # Adding data
        for row in data:
            sheet.append(row)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        workbook.save(response)
        return response

    def get_users_data(self):
        users = User.objects.all()
        data = [['Username', 'Name', 'Email']]
        for user in users:
            data.append([user.username, user.name, user.email])
        return data

    def get_clientes_data(self):
        clientes = Cliente.objects.all()
        data = [['Username', 'Rut', 'Sexo']]
        for cliente in clientes:
            data.append([cliente.user.username, cliente.rut, cliente.sexo])
        return data

    def get_boletas_data(self):
        boletas = Boleta.objects.all()
        data = [['fecha_emision', 'Cliente', 'Cliente Username', 'Cliente Rut', 'Agenda', 'Dia']]
        for boleta in boletas:
            cliente_data = f"{boleta.cliente.user.username} ({boleta.cliente.rut})"
            cliente_username = boleta.cliente.user.username
            cliente_rut = boleta.cliente.rut
            agenda_info = f"Cancha {boleta.reserva.agenda.cancha.numeracion}, {boleta.reserva.agenda.horario}"
            dia = boleta.reserva.dia
            data.append([boleta.fecha_emision, cliente_data, cliente_username, cliente_rut, agenda_info, dia])
        return data


    def get_reservas_data(self):
        reservas = Reserva.objects.all()
        data = [['Cliente', 'Cancha', 'Horario', 'Dia']]
        for reserva in reservas:
            cliente_info = f"{reserva.cliente.user.username} ({reserva.cliente.rut})"
            cancha_info = f"Cancha {reserva.agenda.cancha.numeracion}"
            horario_info = f"{reserva.agenda.horario}"
            data.append([cliente_info, cancha_info, horario_info, reserva.dia])
        return data

    def get(self, request):
        users_data = self.get_users_data()
        clientes_data = self.get_clientes_data()
        boletas_data = self.get_boletas_data()
        reservas_data = self.get_reservas_data()

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=excel_report.xlsx'

        workbook = Workbook()
        users_sheet = workbook.active
        users_sheet.title = 'Users Report'
        for row in users_data:
            users_sheet.append(row)

        clientes_sheet = workbook.create_sheet(title='Clientes Report')
        for row in clientes_data:
            clientes_sheet.append(row)

        boletas_sheet = workbook.create_sheet(title='Boletas Report')
        for row in boletas_data:
            boletas_sheet.append(row)

        reservas_sheet = workbook.create_sheet(title='Reservas Report')
        for row in reservas_data:
            reservas_sheet.append(row)

        workbook.save(response)
        return response

#FUNCIONES PARA SEGUNDO PAGO VALIDACION MOVIL (TRABAJADOR)
class ReservaPorCodigoAPIView(APIView):
    parser_classes = [JSONParser]
    def post(self, request, format=None):
        try:
            codigo = request.data.get('codigo')
            if codigo is None:
                raise ParseError('Campo "codigo" es requerido en el cuerpo JSON.')
            try:
                ticket = Ticket.objects.get(codigo=codigo)
                reserva = ticket.reserva
                reserva.estado = 'finalizada'
                reserva.save()

                boleta_anterior = Boleta.objects.get(reserva=reserva)
                nuevo_pago = boleta_anterior.total
                nueva_boleta = Boleta.objects.create(cliente=boleta_anterior.cliente, reserva=reserva, total=nuevo_pago)
                    
                # boleta = Boleta.objects.get(reserva=reserva)
                # otro_pago = boleta.total 
                # boleta.total += otro_pago
                # boleta.save()
                
                return Response({'Reserva pagada': str(reserva)}, status=status.HTTP_200_OK)
            except Ticket.DoesNotExist:
                return Response({'error': 'Ticket no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        






#API PARA LISTAR TODOS LOS COMPLEJOS
class ComplejoDeportivoListView(View):
    def get(self, request, *args, **kwargs):
        depo = ComplejoDeportivo.objects.all()
        depo_data = serializers.serialize('json', depo)
        return JsonResponse({'depo': depo_data}, safe=False)



class ReservasPorUsuario(ListAPIView):
    serializer_class = ReservaSerializer

    def get_queryset(self):
        username = self.kwargs['username']
        try:
            cliente = Cliente.objects.get(user__username=username)
            return Reserva.objects.filter(cliente=cliente)
        except Cliente.DoesNotExist:
            return Reserva.objects.none()
